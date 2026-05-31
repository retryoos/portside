"""Snapshot tests for the Excel laytime export (notes/architecture_weeks_5_to_8.md §1.1).

Locks the workbook shape so the public-API consumer can rely on it: sheet
names, sheet order, and the canonical Summary!B7 = 84375.0 quantum.
"""

from __future__ import annotations

import io
import sys
from pathlib import Path

import pytest
from openpyxl import load_workbook

_API_ROOT = Path(__file__).resolve().parents[1]
if str(_API_ROOT) not in sys.path:
    sys.path.insert(0, str(_API_ROOT))


# pylint: disable=wrong-import-position
from portside_api.exports import excel as excel_export  # noqa: E402
from portside_api.fixtures import demo_voyage_fixture  # noqa: E402
from portside_api.schemas import LaytimeResult, VoyageState  # noqa: E402


@pytest.fixture
def rotterdam_state() -> VoyageState:
    return demo_voyage_fixture("v_test", "owner")


def _open(workbook_bytes: bytes):
    return load_workbook(io.BytesIO(workbook_bytes), data_only=True)


def test_three_sheets_in_expected_order(rotterdam_state: VoyageState) -> None:
    wb = _open(excel_export.render_laytime_workbook(rotterdam_state))

    # Calculation must be first so it is what opens by default; Summary and
    # Letter follow.
    assert wb.sheetnames == ["Calculation", "Summary", "Letter"]


def test_summary_quantum_locked_to_rotterdam_84375(
    rotterdam_state: VoyageState,
) -> None:
    wb = _open(excel_export.render_laytime_workbook(rotterdam_state))
    summary = wb["Summary"]

    # The canonical cell. If this number ever changes from 84375.0 without a
    # schema announcement the gate breaks.
    assert summary["B7"].value == 84375.0
    # The label next to it must say what the number is.
    assert summary["A7"].value == "Demurrage due"


def test_summary_carries_the_other_headline_figures(
    rotterdam_state: VoyageState,
) -> None:
    wb = _open(excel_export.render_laytime_workbook(rotterdam_state))
    summary = wb["Summary"]
    lt: LaytimeResult = rotterdam_state.laytime  # type: ignore[assignment]

    assert summary["B3"].value == round(float(lt.laytime_allowed_hours), 2)
    assert summary["B4"].value == round(float(lt.laytime_used_hours), 2)
    assert summary["B5"].value == round(float(lt.time_on_demurrage_hours), 2)
    assert summary["B6"].value == round(float(lt.demurrage_rate_per_hour_eur), 2)


def test_calculation_sheet_one_row_per_laytime_row(
    rotterdam_state: VoyageState,
) -> None:
    wb = _open(excel_export.render_laytime_workbook(rotterdam_state))
    calc = wb["Calculation"]

    # Header is row 1; the body rows match laytime.rows 1:1.
    assert calc["A1"].value == "Timestamp"
    assert calc["D1"].value == "Duration h"
    body_rows = list(calc.iter_rows(min_row=2, values_only=True))
    assert len(body_rows) == len(rotterdam_state.laytime.rows)  # type: ignore[union-attr]


def test_contested_row_marks_status_column(rotterdam_state: VoyageState) -> None:
    wb = _open(excel_export.render_laytime_workbook(rotterdam_state))
    calc = wb["Calculation"]

    # The Rotterdam fixture contests exactly the weather stoppage rows. Status
    # column (F) on each of those rows reads "Contested".
    contested_row_indexes = [
        i for i, r in enumerate(rotterdam_state.laytime.rows)  # type: ignore[union-attr]
        if r.contestable
    ]
    assert contested_row_indexes, "Rotterdam fixture should have ≥1 contested row"
    for body_idx in contested_row_indexes:
        sheet_row = body_idx + 2  # +1 for header, +1 for 1-indexed openpyxl
        assert calc.cell(row=sheet_row, column=6).value == "Contested"


def test_letter_sheet_contains_letter_text(rotterdam_state: VoyageState) -> None:
    wb = _open(excel_export.render_laytime_workbook(rotterdam_state))
    letter = wb["Letter"]

    # The letter sheet should at minimum carry the headline EUR figure as
    # plain text so a lawyer scrolling through the workbook still sees it.
    all_text = " ".join(
        str(cell.value) for cell in letter["A"] if cell.value is not None
    )
    assert "84,375.00" in all_text


def test_render_raises_when_pipeline_incomplete() -> None:
    """A voyage that has no laytime yet is a 409 in the route; the renderer
    raises ValueError so the route can translate it."""
    incomplete = VoyageState(
        voyage_id="v_partial", perspective="owner", stage="extracting"
    )
    with pytest.raises(ValueError):
        excel_export.render_laytime_workbook(incomplete)
