"""Excel laytime export (notes/architecture_weeks_5_to_8.md §1.1).

Deterministic, three-sheet `.xlsx` artifact rendered server-side from a
completed ``VoyageState``. The shape is fixed so the route can be hit by the
public API (per notes/product_roadmap.md §2.5) and the workbook stays stable
under snapshot tests.

Sheets:

- ``Calculation``: the laytime ledger from ``LaytimeResult.rows``. Columns:
  Timestamp, Description, Category, Duration h, Cum h, Status. Contested rows
  fill the warning-container amber that the UI uses for contested SoF rows.
- ``Summary``: the key/value strip from the UI's LaytimeSummary. The total
  cell (``B7``) is the canonical quantum; the Rotterdam fixture locks it to
  ``84375.0``.
- ``Letter``: a flat dump of ``packet.claim_letter_markdown`` rendered as
  plain text so a lawyer who already opened the file has the letter alongside
  the math.

The model never owns a cell. Every value comes from a deterministic field on
``VoyageState``; the ``claim_letter_markdown`` is the only string of model
output that hits the workbook, and it lands on a sheet of its own so a glitch
inside the prose cannot break a numeric assertion.
"""

from __future__ import annotations

import io
from datetime import datetime
from typing import TYPE_CHECKING

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

if TYPE_CHECKING:
    from ..schemas import VoyageState


def _naive(dt: datetime) -> datetime:
    """Strip tzinfo: openpyxl rejects timezone-aware datetimes. Voyage data is
    UTC throughout, so dropping the marker (rather than converting) is correct
    and keeps the cell's wall time identical to the API JSON."""
    return dt.replace(tzinfo=None) if dt.tzinfo is not None else dt


# Amber container fill, matching the UI's `bg-contested-container` token.
_CONTESTED_FILL = PatternFill(start_color="FFFFECD1", end_color="FFFFECD1", fill_type="solid")
_HEADER_FILL = PatternFill(start_color="FFF3F4F5", end_color="FFF3F4F5", fill_type="solid")
_HEADER_FONT = Font(bold=True, size=10, color="FF6D6E72")
_BODY_FONT = Font(size=10)
_TOTAL_FONT = Font(bold=True, size=11)

_DATE_FMT = "yyyy-mm-dd hh:mm"
_HOURS_FMT = "0.00"
_CURRENCY_FMT = '"EUR" #,##0.00'

_CATEGORY_LABEL = {
    "laytime": "Laytime",
    "demurrage": "Demurrage",
    "excepted": "Excepted",
}


def render_laytime_workbook(state: "VoyageState") -> bytes:
    """Render a `VoyageState` to an `.xlsx` byte string.

    Raises ``ValueError`` if the pipeline has not produced enough data to render
    a meaningful workbook (no extraction or no laytime). The route layer turns
    that into a 409.
    """
    if state.extraction is None or state.laytime is None:
        raise ValueError("voyage is not ready: extraction and laytime are required")

    wb = Workbook()
    # `Workbook()` ships a default sheet named "Sheet"; rename and reuse it.
    summary_ws = wb.active
    summary_ws.title = "Summary"
    _write_summary(summary_ws, state)

    calc_ws = wb.create_sheet("Calculation")
    _write_calculation(calc_ws, state)

    letter_ws = wb.create_sheet("Letter")
    _write_letter(letter_ws, state)

    # `Sheet order` matters in Excel; put Calculation first since that is what
    # the recipient lands on by default and what the lawyer audits.
    wb.move_sheet("Calculation", offset=-1)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _write_summary(ws, state: "VoyageState") -> None:
    """Two-column key/value strip + the headline quantum at B7.

    Layout is fixed so the snapshot test can assert specific cells:
        A1: "Calculation summary"
        A2: vessel name
        A3..B5: laytime allowed / used / on demurrage
        A6:    "Demurrage rate" -> B6 currency
        A7:    "Demurrage due"  -> B7 currency  (the canonical quantum)
        A8:    "Time bar date"  -> B8 ISO date
    """
    cp = state.extraction.charter_party  # type: ignore[union-attr]
    lt = state.laytime  # type: ignore[union-attr]
    packet = state.packet

    ws["A1"] = "Calculation summary"
    ws["A1"].font = Font(bold=True, size=14)

    ws["A2"] = "Vessel"
    ws["B2"] = cp.vessel_name
    ws["B2"].font = _BODY_FONT

    ws["A3"] = "Laytime allowed (h)"
    ws["B3"] = round(float(lt.laytime_allowed_hours), 2)
    ws["B3"].number_format = _HOURS_FMT

    ws["A4"] = "Laytime used (h)"
    ws["B4"] = round(float(lt.laytime_used_hours), 2)
    ws["B4"].number_format = _HOURS_FMT

    ws["A5"] = "On demurrage (h)"
    ws["B5"] = round(float(lt.time_on_demurrage_hours), 2)
    ws["B5"].number_format = _HOURS_FMT

    ws["A6"] = "Demurrage rate / h"
    ws["B6"] = round(float(lt.demurrage_rate_per_hour_eur), 2)
    ws["B6"].number_format = _CURRENCY_FMT

    ws["A7"] = "Demurrage due"
    ws["B7"] = round(float(lt.demurrage_due_eur), 2)
    ws["B7"].number_format = _CURRENCY_FMT
    ws["A7"].font = _TOTAL_FONT
    ws["B7"].font = _TOTAL_FONT

    ws["A8"] = "Time bar date"
    ws["B8"] = packet.time_bar_date if packet else ""

    ws["A9"] = "Submitted within time bar"
    ws["B9"] = "Yes" if packet and packet.submitted_within_time_bar else "No"

    # Column widths chosen for readability without auto-fit (auto-fit varies by
    # client; fixed widths render the same across Excel, Google Sheets, Numbers).
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 22


def _write_calculation(ws, state: "VoyageState") -> None:
    """The laytime ledger. One row per ``LaytimeRow``; contested rows amber.

    Header row sits at row 1 with a muted fill; the body starts at row 2.
    """
    headers = ["Timestamp", "Description", "Category", "Duration h", "Cum h", "Status"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT

    for r_idx, row in enumerate(state.laytime.rows, start=2):  # type: ignore[union-attr]
        ws.cell(row=r_idx, column=1, value=_naive(row.from_ts)).number_format = _DATE_FMT
        ws.cell(row=r_idx, column=2, value=row.reason).font = _BODY_FONT
        ws.cell(
            row=r_idx, column=3, value=_CATEGORY_LABEL.get(row.status, row.status)
        ).font = _BODY_FONT
        ws.cell(row=r_idx, column=4, value=round(float(row.duration_hours), 2)).number_format = _HOURS_FMT
        ws.cell(row=r_idx, column=5, value=round(float(row.running_total_hours), 2)).number_format = _HOURS_FMT
        ws.cell(
            row=r_idx,
            column=6,
            value="Contested" if row.contestable else "",
        ).font = _BODY_FONT

        if row.contestable:
            for col_idx in range(1, len(headers) + 1):
                ws.cell(row=r_idx, column=col_idx).fill = _CONTESTED_FILL

    # Right-align the numeric columns.
    for col in (4, 5):
        letter = get_column_letter(col)
        ws.column_dimensions[letter].width = 14
        for cell in ws[letter][1:]:
            cell.alignment = Alignment(horizontal="right")

    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 60
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["F"].width = 14

    # Freeze the header so a long SoF stays scannable.
    ws.freeze_panes = "A2"


def _write_letter(ws, state: "VoyageState") -> None:
    """Letter sheet: flat text dump of ``packet.claim_letter_markdown``.

    Each markdown line lands in column A. No formatting beyond Inter-equivalent
    body font; we are not reproducing the rendered HTML, just giving the lawyer
    text to read alongside the math.
    """
    ws.column_dimensions["A"].width = 110
    text = state.packet.claim_letter_markdown if state.packet else ""
    for r_idx, line in enumerate(text.splitlines() or [""], start=1):
        cell = ws.cell(row=r_idx, column=1, value=line)
        cell.font = _BODY_FONT
        cell.alignment = Alignment(wrap_text=False, vertical="top")
